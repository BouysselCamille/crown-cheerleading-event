#!/usr/bin/env python3
"""Génère le fichier Excel d'inscription Crown Cheerleading Events."""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, Protection
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

# ── Couleurs ──────────────────────────────────────────────────────────────────
GOLD        = "D4AF37"
GOLD_DARK   = "A0820A"
BLACK       = "1A1A1A"
DARK        = "FFFFFF"
GRAY        = "F9F9F9"
GRAY2       = "F0F0F0"
WHITE       = "1A1A1A"
LIGHT_GRAY  = "111111"
MID_GRAY    = "555555"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=11, italic=False, name="Calibri"):
    return Font(bold=bold, color=color, size=size, italic=italic, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color=GOLD_DARK)
    return Border(top=s, bottom=s, left=s, right=s)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def merge_style(ws, cell_range, value, bg, fg=WHITE, size=11, bold=False,
                h="left", wrap=False, locked=True):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = value
    c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size)
    c.alignment = align(h=h, wrap=wrap)
    c.protection = Protection(locked=locked)

def header_cell(ws, row, col, value, bg=GRAY, fg=GOLD, size=9, bold=True, h="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(bold=bold, color=fg, size=size)
    c.alignment = align(h=h, v="center", wrap=True)
    c.border = thin_border()
    c.protection = Protection(locked=True)
    return c

def label_cell(ws, row, col, value, bg=GRAY, fg=MID_GRAY, size=9):
    c = ws.cell(row=row, column=col, value=value)
    c.fill = fill(bg)
    c.font = font(color=fg, size=size)
    c.alignment = align()
    c.protection = Protection(locked=True)
    return c

def input_cell(ws, row, col, bg=GRAY2, fg=WHITE):
    """Cellule éditable par l'utilisateur."""
    c = ws.cell(row=row, column=col)
    c.fill = fill(bg)
    c.font = font(color=fg, size=10)
    c.alignment = align()
    c.border = Border(bottom=Side(style="thin", color=GOLD_DARK))
    c.protection = Protection(locked=False)
    return c

def calc_cell(ws, row, col, formula, bg=DARK, fg=GOLD, size=11, bold=False, h="center"):
    """Cellule calculée (formule), verrouillée."""
    c = ws.cell(row=row, column=col, value=formula)
    c.fill = fill(bg)
    c.font = font(color=fg, size=size, bold=bold)
    c.alignment = align(h=h)
    c.border = Border(bottom=Side(style="thin", color=GOLD_DARK))
    c.protection = Protection(locked=True)
    return c

# ── WORKBOOK ──────────────────────────────────────────────────────────────────
wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 1 — Accueil
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Accueil"
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = GOLD

# Colonnes : A(marge 3), B(label 24), C(Type 18), D(sep 4), E(Âge 14), F(sep 4), G(Niveau 16)
for col, w in [(1,3),(2,24),(3,18),(4,4),(5,14),(6,4),(7,16)]:
    set_col_width(ws1, col, w)

# ── Ligne 1 : Header ──────────────────────────────────────────────────────────
ws1.row_dimensions[1].height = 50
ws1.merge_cells("A1:G1")
c = ws1["A1"]
c.value = "CROWN CHEERLEADING EVENTS — DOSSIER D'INSCRIPTION"
c.fill = fill(BLACK)
c.font = Font(bold=True, color=GOLD, size=16, name="Calibri")
c.alignment = align(h="center", v="center")
c.protection = Protection(locked=True)

# ── Ligne 2 : Sous-titre ──────────────────────────────────────────────────────
ws1.row_dimensions[2].height = 22
ws1.merge_cells("A2:G2")
c = ws1["A2"]
c.value = "4 avril 2027  ·  Dojo de Paris  ·  Un fichier par équipe"
c.fill = fill(DARK)
c.font = font(color=MID_GRAY, size=9, italic=True)
c.alignment = align(h="center")
c.protection = Protection(locked=True)

ws1.row_dimensions[3].height = 8

# ── Ligne 4 : Section INFORMATIONS GÉNÉRALES ──────────────────────────────────
ws1.row_dimensions[4].height = 22
merge_style(ws1, "A4:G4", "  INFORMATIONS GÉNÉRALES", GOLD_DARK, BLACK, size=10, bold=True)

# ── Lignes 5-6 : Nom équipe, Nom du club ──────────────────────────────────────
for row, label in [(5, "Nom de l'équipe"), (6, "Nom du club")]:
    ws1.row_dimensions[row].height = 22
    label_cell(ws1, row, 2, label, fg=LIGHT_GRAY, size=10)
    input_cell(ws1, row, 3)
    ws1.merge_cells(f"C{row}:G{row}")

# ── Ligne 7 : spacer ─────────────────────────────────────────────────────────
ws1.row_dimensions[7].height = 8

# ── Ligne 8 : Division — liste déroulante unique ──────────────────────────────
ws1.row_dimensions[8].height = 26
label_cell(ws1, 8, 2, "Division", fg=LIGHT_GRAY, size=10)
cd = input_cell(ws1, 8, 3)
cd.alignment = align(h="left")
ws1.merge_cells("C8:G8")
dv_div = DataValidation(type="list",
    formula1="Listes!$A$1:$A$45",
    showDropDown=False, showErrorMessage=True,
    error="Choisissez une division dans la liste", errorTitle="Valeur invalide")
ws1.add_data_validation(dv_div)
dv_div.add(cd)

ws1.row_dimensions[9].height = 10

# ── Ligne 10 : Section RÉSUMÉ & CALCUL ───────────────────────────────────────
ws1.row_dimensions[10].height = 22
merge_style(ws1, "A10:G10", "  RÉSUMÉ & CALCUL DES FRAIS", GOLD_DARK, BLACK, size=10, bold=True)

summary_rows = [
    (11, "Nombre total d'athlètes",             '=COUNTA(Athlètes!B5:B39)'),
    (12, "Dont athlètes masculins",              '=COUNTIF(Athlètes!E5:E200,"Masculin")'),
    (13, "Accompagnateurs inclus (2/équipe)",    2),
    (14, "Accompagnateurs supplémentaires",      '=MAX(0,COUNTA(Accompagnateurs!B5:B24)-2)'),
    (15, "Coût accompagnateurs suppl. (€)",      '=C14*40'),
]
for row, label, value in summary_rows:
    ws1.row_dimensions[row].height = 22
    label_cell(ws1, row, 2, label, fg=LIGHT_GRAY, size=10)
    calc_cell(ws1, row, 3, value)

ws1.row_dimensions[16].height = 10

# ── Ligne 17 : Prix par athlète (calculé automatiquement selon le type) ───────
ws1.row_dimensions[17].height = 26
merge_style(ws1, "A17:B17", "  Prix par athlète (€)", GRAY, MID_GRAY, size=9)
cprix = ws1["C17"]
cprix.value = '=IF(C8="Démo",37.5,IF(ISNUMBER(SEARCH("Novice",C8)),40,IF(OR(ISNUMBER(SEARCH("Allstar",C8)),ISNUMBER(SEARCH("Prep",C8)),ISNUMBER(SEARCH("Universitaire",C8))),50,IF(C8<>"","50",""))))'
cprix.fill = fill(GRAY2)
cprix.font = font(color=GOLD_DARK, size=11, bold=True)
cprix.alignment = align(h="center")
cprix.border = Border(bottom=Side(style="thin", color=GOLD_DARK))
cprix.protection = Protection(locked=True)

cnote17 = ws1.cell(row=17, column=4,
    value='=IF(C8="Démo","Démo : 37,50 €",IF(ISNUMBER(SEARCH("Novice",C8)),"Novice : 40 €",IF(OR(ISNUMBER(SEARCH("Allstar",C8)),ISNUMBER(SEARCH("Prep",C8)),ISNUMBER(SEARCH("Universitaire",C8))),"AllStar / Prep / Univ. : 50 €",IF(C8<>"","50 €","← sélectionnez une division"))))')
cnote17.fill = fill(GRAY2)
cnote17.font = font(color=MID_GRAY, size=8, italic=True)
cnote17.alignment = align(h="left")
cnote17.protection = Protection(locked=True)
ws1.merge_cells("D17:G17")

ws1.row_dimensions[18].height = 10

# ── Ligne 19 : TOTAL ESTIMÉ ──────────────────────────────────────────────────
ws1.row_dimensions[19].height = 32
merge_style(ws1, "A19:B19", "  TOTAL ESTIMÉ (athlètes + accomp. suppl.)", BLACK, GOLD, size=10, bold=True)
ct = ws1["C19"]
ct.value = "=C11*C17+C15"
ct.fill = fill(BLACK)
ct.font = Font(bold=True, color=GOLD, size=14, name="Calibri")
ct.alignment = align(h="center")
ct.border = Border(top=Side(style="medium", color=GOLD), bottom=Side(style="medium", color=GOLD))
ct.protection = Protection(locked=True)

ws1.row_dimensions[20].height = 10

# ── Lignes 21-22 : Notes ──────────────────────────────────────────────────────
ws1.row_dimensions[21].height = 36
merge_style(ws1, "A21:G21",
    "Ce document ne fait pas office de facture. Les tarifs affichés sur le site "
    "crown-cheerleading-events.fr font valeur de référence.",
    GOLD_DARK, BLACK, size=9, wrap=True, bold=False)

ws1.row_dimensions[22].height = 50
merge_style(ws1, "A22:G22",
    "Un acompte de 30 % est requis à l'inscription. "
    "Toute annulation avant la fermeture des inscriptions donne droit à un remboursement de 70 %, "
    "les 30 % d'acompte restant acquis. Envoyez ce fichier à : contact@crown-cheerleading-events.fr",
    DARK, MID_GRAY, size=8, wrap=True)

# ── Protection feuille Accueil ────────────────────────────────────────────────
ws1.protection.sheet = True
ws1.protection.password = ""
ws1.protection.enable()


# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 2 — Athlètes
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Athlètes")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = GOLD_DARK

for col, w in [(1,5),(2,22),(3,22),(4,18),(5,14),(6,16)]:
    set_col_width(ws2, col, w)

ws2.row_dimensions[1].height = 40
ws2.merge_cells("A1:F1")
c = ws2["A1"]
c.value = "LISTE DES ATHLÈTES"
c.fill = fill(BLACK)
c.font = Font(bold=True, color=GOLD, size=14, name="Calibri")
c.alignment = align(h="center", v="center")
c.protection = Protection(locked=True)

ws2.row_dimensions[2].height = 14
ws2.merge_cells("A2:F2")
ws2["A2"].fill = fill(DARK)
ws2["A2"].protection = Protection(locked=True)

ws2.row_dimensions[3].height = 20
merge_style(ws2, "A3:F3", '← Équipe : voir feuille "Accueil"', GRAY, MID_GRAY, size=8)

ws2.row_dimensions[4].height = 28
for col, label in [(1,"#"),(2,"Nom"),(3,"Prénom"),(4,"Date de naissance\n(JJ/MM/AAAA)"),(5,"Sexe"),(6,"Remarque")]:
    header_cell(ws2, 4, col, label, bg=GRAY2, fg=GOLD, size=9)

dv_sex = DataValidation(type="list", formula1='"Féminin,Masculin"', showDropDown=False)
ws2.add_data_validation(dv_sex)

for row in range(5, 40):
    ws2.row_dimensions[row].height = 20
    bg = DARK if row % 2 == 0 else GRAY
    c = ws2.cell(row=row, column=1, value=row - 4)
    c.fill = fill(bg)
    c.font = font(color=MID_GRAY, size=9)
    c.alignment = align(h="center")
    c.protection = Protection(locked=True)
    for col in [2, 3]:
        ci = ws2.cell(row=row, column=col)
        ci.fill = fill(bg)
        ci.font = font(color=WHITE, size=10)
        ci.alignment = align()
        ci.border = Border(bottom=Side(style="hair", color=GRAY2))
        ci.protection = Protection(locked=False)
    cd = ws2.cell(row=row, column=4)
    cd.fill = fill(bg)
    cd.font = font(color=WHITE, size=10)
    cd.alignment = align(h="center")
    cd.number_format = "DD/MM/YYYY"
    cd.border = Border(bottom=Side(style="hair", color=GRAY2))
    cd.protection = Protection(locked=False)
    cs = ws2.cell(row=row, column=5)
    cs.fill = fill(bg)
    cs.font = font(color=GOLD, size=10)
    cs.alignment = align(h="center")
    cs.protection = Protection(locked=False)
    dv_sex.add(cs)
    cr = ws2.cell(row=row, column=6)
    cr.fill = fill(bg)
    cr.font = font(color=MID_GRAY, size=9, italic=True)
    cr.alignment = align()
    cr.protection = Protection(locked=False)

ws2.protection.sheet = True
ws2.protection.password = ""
ws2.protection.enable()


# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE 3 — Accompagnateurs
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Accompagnateurs")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = GOLD_DARK

for col, w in [(1,5),(2,22),(3,22),(4,32),(5,18)]:
    set_col_width(ws3, col, w)

ws3.row_dimensions[1].height = 40
ws3.merge_cells("A1:E1")
c = ws3["A1"]
c.value = "ENTRAÎNEURS & ACCOMPAGNATEURS"
c.fill = fill(BLACK)
c.font = Font(bold=True, color=GOLD, size=14, name="Calibri")
c.alignment = align(h="center", v="center")
c.protection = Protection(locked=True)

ws3.row_dimensions[2].height = 14
ws3.merge_cells("A2:E2")
ws3["A2"].fill = fill(DARK)
ws3["A2"].protection = Protection(locked=True)

ws3.row_dimensions[3].height = 30
merge_style(ws3, "A3:E3",
    "2 accompagnateurs par équipe sont inclus dans les frais d'inscription (lignes 1 et 2). "
    "Les suivants sont facturés 40 € chacun et comptabilisés automatiquement dans l'onglet Accueil.",
    GRAY, MID_GRAY, size=8, wrap=True)

ws3.row_dimensions[4].height = 28
for col, label in [(1,"#"),(2,"Nom"),(3,"Prénom"),(4,"Email (obligatoire pour entraîneur)"),(5,"Rôle")]:
    header_cell(ws3, 4, col, label, bg=GRAY2, fg=GOLD, size=9)

dv_role = DataValidation(
    type="list",
    formula1='"Entraîneur principal,Assistant entraîneur,Parent accompagnateur,Physio / Thérapeute,Photographe / Média,Autre"',
    showDropDown=False
)
ws3.add_data_validation(dv_role)

for row in range(5, 25):
    ws3.row_dimensions[row].height = 22
    bg = DARK if row % 2 == 0 else GRAY
    idx = row - 4
    highlight = idx <= 2
    c = ws3.cell(row=row, column=1, value=f"{idx} \u2713" if highlight else idx)
    c.fill = fill(GOLD_DARK if highlight else bg)
    c.font = font(color=BLACK if highlight else MID_GRAY, size=9, bold=highlight)
    c.alignment = align(h="center")
    c.protection = Protection(locked=True)
    for col in [2, 3, 4]:
        ci = ws3.cell(row=row, column=col)
        ci.fill = fill(GRAY2 if highlight else bg)
        ci.font = font(color=WHITE, size=10)
        ci.alignment = align()
        ci.border = Border(bottom=Side(style="hair", color=GRAY2))
        ci.protection = Protection(locked=False)
    cr = ws3.cell(row=row, column=5)
    cr.fill = fill(GRAY2 if highlight else bg)
    cr.font = font(color=GOLD if highlight else LIGHT_GRAY, size=10)
    cr.alignment = align(h="center")
    cr.protection = Protection(locked=False)
    dv_role.add(cr)

ws3.row_dimensions[25].height = 10
ws3.row_dimensions[26].height = 50
merge_style(ws3, "A26:E26",
    "RAPPEL RÈGLEMENT : Seuls les accompagnateurs et athlètes déclarés sont autorisés en salle "
    "d'échauffement et dans les vestiaires. Les accompagnateurs doivent arriver en même temps que leur équipe "
    "et se présenter à l'accueil pour récupérer leur bracelet.",
    DARK, MID_GRAY, size=8, wrap=True)

ws3.protection.sheet = True
ws3.protection.password = ""
ws3.protection.enable()

# ══════════════════════════════════════════════════════════════════════════════
# FEUILLE CACHÉE — Listes (source des dropdowns)
# ══════════════════════════════════════════════════════════════════════════════
ws_lists = wb.create_sheet("Listes")
ws_lists.sheet_state = "hidden"

DIVISIONS = [
    "Démo",
    "U6 Novice L1",
    "U8 Novice L1",
    "U12 Prep L1",
    "U12 Prep L2.1",
    "U16 Prep L1",
    "U16 Prep L2.1",
    "U18 Prep L1",
    "U18 Prep L2.1",
    "Open Prep L1",
    "Open Prep L2.1",
    "U12 Allstar L1",
    "U12 Allstar L2",
    "U16 Allstar L1",
    "U16 Allstar L2",
    "U16 Allstar L3",
    "U18 Allstar L1",
    "U18 Allstar L2",
    "U18 Allstar L3",
    "U18 Allstar L4",
    "Open Allstar L1",
    "Open Allstar L2",
    "Open AG Allstar L3",
    "Open AG Allstar L4",
    "Open AG Allstar L5",
    "Open AG Allstar L6",
    "Open AG Allstar L7",
    "Open Coed Allstar L3",
    "Open Coed Allstar L4",
    "Open Coed Allstar L5",
    "Open Coed Allstar L6",
    "Open Coed Allstar L7",
    "Open NT Allstar L2",
    "Open AG NT Allstar L3",
    "Open AG NT Allstar L4",
    "Open AG NT Allstar L5",
    "Open AG NT Allstar L6",
    "Open AG NT Allstar L7",
    "Open Coed NT Allstar L3",
    "Open Coed NT Allstar L4",
    "Open Coed NT Allstar L5",
    "Open Coed NT Allstar L6",
    "Open Coed NT Allstar L7",
    "Universitaire L2",
    "Universitaire L3",
]

for i, div in enumerate(DIVISIONS, start=1):
    ws_lists.cell(row=i, column=1, value=div)

# ── Sauvegarde ────────────────────────────────────────────────────────────────
output = "/Users/camille.bouyssel/Cheer/crown-cheerleading-event/public/docs/inscription-equipe.xlsx"
wb.save(output)
print(f"Fichier cree : {output}")
